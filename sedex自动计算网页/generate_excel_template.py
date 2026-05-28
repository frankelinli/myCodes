from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


THIN = Side(style="thin", color="B0BEC5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FONT_NORMAL = Font(name="Microsoft YaHei", size=11, color="222222")
FONT_BOLD = Font(name="Microsoft YaHei", size=11, bold=True, color="1A3A5C")
FONT_TITLE = Font(name="Microsoft YaHei", size=18, bold=True, color="1A3A5C")
FONT_SUBTITLE = Font(name="Microsoft YaHei", size=10, color="777777")
FONT_SECTION = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
FONT_GREEN = Font(name="Microsoft YaHei", size=11, bold=True, color="1E6B2E")

FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
FILL_TABLE_HEAD = PatternFill("solid", fgColor="DCE6F1")
FILL_LABEL = PatternFill("solid", fgColor="F5F8FF")
FILL_CALC = PatternFill("solid", fgColor="D9D9D9")
FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FILL_SECTION = PatternFill("solid", fgColor="1A3A5C")
FILL_NOTICE = PatternFill("solid", fgColor="FFFBE6")
FILL_NOTE = PatternFill("solid", fgColor="F5F7FA")


def style_cell(cell, *, font=FONT_NORMAL, fill=FILL_WHITE, align=None, border=BORDER, number_format=None):
    cell.font = font
    cell.fill = fill
    cell.border = border
    cell.alignment = align or Alignment(horizontal="center", vertical="center", wrap_text=True)
    if number_format:
        cell.number_format = number_format


def merge_write(ws, cell_range, value, *, font=FONT_NORMAL, fill=FILL_WHITE, align=None, border=BORDER, number_format=None):
    ws.merge_cells(cell_range)
    start = ws[cell_range.split(":")[0]]
    start.value = value
    min_col, min_row, max_col, max_row = ws[cell_range][0][0].column, ws[cell_range][0][0].row, ws[cell_range][-1][-1].column, ws[cell_range][-1][-1].row
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            style_cell(cell, font=font, fill=fill, align=align, border=border, number_format=number_format)
    return start


def write_row(ws, row_idx, values, styles=None):
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell_style = styles[col_idx - 1] if styles else {}
        style_cell(cell, **cell_style)


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def formula_pct(numerator, denominator):
    return f'=IF({denominator}>0,{numerator}/{denominator},"")'


def diff_formula(actual, legal, decimals):
    fmt = "0" if decimals == 0 else "0." + ("0" * decimals)
    return f'=IF(ISNUMBER({actual}-{legal}),"实际"&IF({actual}-{legal}>=0,"+","")&TEXT({actual}-{legal},"{fmt}"),"-")'


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sedex审核信息收集表"
    ws.freeze_panes = "A9"
    set_widths(
        ws,
        {
            "A": 10,
            "B": 28,
            "C": 11,
            "D": 11,
            "E": 11,
            "F": 14,
            "G": 14,
            "H": 18,
            "I": 4,
            "J": 20,
        },
    )
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20
    ws.sheet_view.showGridLines = False

    merge_write(ws, "A1:H1", "Sedex 审核信息收集表", font=FONT_TITLE, fill=FILL_WHITE, align=Alignment(horizontal="center", vertical="center"), border=Border())
    merge_write(ws, "A2:H2", "Social Audit Information Collection Form", font=FONT_SUBTITLE, fill=FILL_WHITE, align=Alignment(horizontal="center", vertical="center"), border=Border())

    merge_write(ws, "A4:B4", "工厂名称", font=FONT_BOLD, align=Alignment(horizontal="left", vertical="center"), border=Border())
    merge_write(ws, "C4:D4", "", fill=FILL_WHITE, align=Alignment(horizontal="left", vertical="center"))
    merge_write(ws, "E4:F4", "填表日期", font=FONT_BOLD, align=Alignment(horizontal="left", vertical="center"), border=Border())
    ws["G4"] = ""
    style_cell(ws["G4"], align=Alignment(horizontal="left", vertical="center"), number_format="yyyy-mm-dd")
    style_cell(ws["H4"], font=FONT_BOLD, fill=FILL_WHITE, align=Alignment(horizontal="left", vertical="center"), border=Border())
    ws["H4"] = "填表人"
    ws["H5"] = ""
    style_cell(ws["H5"], align=Alignment(horizontal="left", vertical="center"))
    merge_write(ws, "A6:H6", "填表说明：灰色区域为自动计算结果，无需手动填写。请在白色/浅色输入框中填入实际数据，表格将实时自动计算相关数据。", fill=FILL_NOTICE, align=Alignment(horizontal="left", vertical="center", wrap_text=True))

    merge_write(ws, "A8:H8", "一、人员信息 / Workforce Information", font=FONT_SECTION, fill=FILL_SECTION, align=Alignment(horizontal="left", vertical="center"))
    workforce_header_styles = [
        {"font": FONT_BOLD, "fill": FILL_TABLE_HEAD},
        {"font": FONT_BOLD, "fill": FILL_TABLE_HEAD},
        {"font": FONT_BOLD, "fill": FILL_TABLE_HEAD},
        {"font": FONT_BOLD, "fill": FILL_TABLE_HEAD},
        {"font": FONT_BOLD, "fill": FILL_TABLE_HEAD},
        {"font": FONT_BOLD, "fill": FILL_TABLE_HEAD},
        {"font": FONT_BOLD, "fill": FILL_TABLE_HEAD},
        {"font": FONT_BOLD, "fill": FILL_TABLE_HEAD},
    ]
    write_row(ws, 9, ["分类", "岗位 / 说明", "男", "女", "合计", "男占比", "女占比", "全厂总人数"], workforce_header_styles)

    merge_write(ws, "A10:A16", "生产", font=FONT_BOLD, fill=FILL_TABLE_HEAD)
    merge_write(ws, "H10:H16", "=E10+E17+E18+E19", font=FONT_GREEN, fill=FILL_GREEN, number_format="0")

    rows = {
        10: ["生产工人", 6, 7, "=C10+D10", formula_pct("C10", "E10"), formula_pct("D10", "E10")],
        11: ["其中：18-24岁工人", 1, 0, "=C11+D11", None, None],
        12: ["其中：最小年龄（岁）", None, None, '=IF(C12>0,C12,"")', None, None],
        13: ["其中：外省工人", 5, 7, "=C13+D13", '=IF(E10>0,"外省占比 "&ROUND(E13/E10*100,0)&"%","外省占比 -")', None],
        14: ["其中：过去12个月入职", 0, 0, "=C14+D14", None, None],
        15: ["其中：达到退休年龄", 0, 0, "=C15+D15", '="男 "&TEXT(C15,"0")&" / 女 "&TEXT(D15,"0")&" / 合计 "&TEXT(E15,"0")', None],
    }

    for row_idx, row_data in rows.items():
        label = ws.cell(row=row_idx, column=2, value=row_data[0])
        style_cell(label, fill=FILL_LABEL, align=Alignment(horizontal="left", vertical="center"))
        if row_idx == 12:
            merge_write(ws, f"C{row_idx}:D{row_idx}", "", align=Alignment(horizontal="left", vertical="center"))
            if row_data[1] is not None:
                ws[f"C{row_idx}"] = row_data[1]
            calc_cell = ws.cell(row=row_idx, column=5, value=row_data[3])
            style_cell(calc_cell, fill=FILL_CALC)
            merge_write(ws, f"F{row_idx}:G{row_idx}", "", fill=FILL_WHITE)
            continue
        for col_idx, value in zip([3, 4, 5], row_data[1:4]):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            fill = FILL_WHITE if col_idx in (3, 4) else FILL_CALC
            number_format = "0" if col_idx != 5 else "0"
            style_cell(cell, fill=fill, number_format=number_format)
        if row_idx == 10:
            for col_idx, formula in zip([6, 7], row_data[4:6]):
                cell = ws.cell(row=row_idx, column=col_idx, value=formula)
                style_cell(cell, fill=FILL_CALC, number_format="0.0%")
        elif row_idx == 11:
            merge_write(ws, f"F{row_idx}:G{row_idx}", "（仅供参考）", fill=FILL_WHITE, font=FONT_SUBTITLE)
        elif row_idx == 13:
            merge_write(ws, f"F{row_idx}:G{row_idx}", row_data[4], fill=FILL_CALC)
        elif row_idx == 14:
            merge_write(ws, f"F{row_idx}:G{row_idx}", "男 / 女 / 合计", fill=FILL_WHITE, font=FONT_SUBTITLE)
        elif row_idx == 15:
            merge_write(ws, f"F{row_idx}:G{row_idx}", row_data[4], fill=FILL_CALC)

    label = ws["B16"]
    label.value = "外省最多的5个省份"
    style_cell(label, fill=FILL_LABEL, align=Alignment(horizontal="left", vertical="center"))
    merge_write(ws, "C16:G16", "Hunan, Hubei, Guizhou, Sichuan, Jiangxi", align=Alignment(horizontal="left", vertical="center"))

    merge_write(ws, "A17:A19", "非生产", font=FONT_BOLD, fill=FILL_TABLE_HEAD)
    non_prod = {
        17: ["高管级别", 1, 0, "=C17+D17"],
        18: ["主管级别", 0, 0, "=C18+D18"],
        19: ["行政、财务、研发、后勤等", 0, 1, "=C19+D19"],
    }
    for row_idx, row_data in non_prod.items():
        style_cell(ws.cell(row=row_idx, column=2, value=row_data[0]), fill=FILL_LABEL, align=Alignment(horizontal="left", vertical="center"))
        style_cell(ws.cell(row=row_idx, column=3, value=row_data[1]), fill=FILL_WHITE)
        style_cell(ws.cell(row=row_idx, column=4, value=row_data[2]), fill=FILL_WHITE)
        style_cell(ws.cell(row=row_idx, column=5, value=row_data[3]), fill=FILL_CALC)
        merge_write(ws, f"F{row_idx}:H{row_idx}", "", fill=FILL_WHITE)

    merge_write(ws, "A22:E22", "二、抽样信息 / Sampling Information", font=FONT_SECTION, fill=FILL_SECTION, align=Alignment(horizontal="left", vertical="center"))
    write_row(
        ws,
        23,
        ["说明", "男", "女", "共抽样人数", "备注"],
        [{"font": FONT_BOLD, "fill": FILL_TABLE_HEAD}] * 5,
    )
    sampling = {
        24: ["抽样人数（合计）", 6, 4, "=B24+C24"],
        25: ["其中：外省工人", 5, 4, "=B25+C25"],
    }
    for row_idx, row_data in sampling.items():
        style_cell(ws.cell(row=row_idx, column=1, value=row_data[0]), fill=FILL_LABEL, align=Alignment(horizontal="left", vertical="center"))
        style_cell(ws.cell(row=row_idx, column=2, value=row_data[1]), fill=FILL_WHITE)
        style_cell(ws.cell(row=row_idx, column=3, value=row_data[2]), fill=FILL_WHITE)
        style_cell(ws.cell(row=row_idx, column=4, value=row_data[3]), fill=FILL_CALC)
        style_cell(ws.cell(row=row_idx, column=5, value=""), fill=FILL_WHITE)

    merge_write(ws, "A28:D28", "三、证件与许可证 / Certificates & Permits", font=FONT_SECTION, fill=FILL_SECTION, align=Alignment(horizontal="left", vertical="center"))
    write_row(
        ws,
        29,
        ["证件名称", "证件 / 报告编号", "颁发日期", "有效期（无固定期限请填长期）"],
        [{"font": FONT_BOLD, "fill": FILL_TABLE_HEAD}] * 4,
    )
    cert_rows = {
        30: ["营业执照", "91440111340203752A", "2015-05-15", "无固定期限"],
        31: ["消防验收报告", "穗公白消（建验）字[2010] 第0124号", "2010-06-21", ""],
        32: ["竣工验收报告", "GD301521", "2010-04-20", ""],
        33: ["其他安全、环保相关证书 1", "", "", ""],
        34: ["其他安全、环保相关证书 2", "", "", ""],
        35: ["其他安全、环保相关证书 3", "", "", ""],
    }
    for row_idx, row_data in cert_rows.items():
        style_cell(ws.cell(row=row_idx, column=1, value=row_data[0]), fill=FILL_LABEL, align=Alignment(horizontal="left", vertical="center"))
        style_cell(ws.cell(row=row_idx, column=2, value=row_data[1]), fill=FILL_WHITE, align=Alignment(horizontal="left", vertical="center"))
        cell_date = ws.cell(row=row_idx, column=3, value=row_data[2])
        style_cell(cell_date, fill=FILL_WHITE, number_format="yyyy-mm-dd")
        style_cell(ws.cell(row=row_idx, column=4, value=row_data[3]), fill=FILL_WHITE, align=Alignment(horizontal="left", vertical="center"))

    for row_idx in range(30, 36):
        ws[f"J{row_idx}"] = (
            f'=IF(AND(B{row_idx}="",C{row_idx}="",D{row_idx}=""),"",'
            f'A{row_idx}&": "&'
            f'IF(B{row_idx}<>"","Doc number: "&B{row_idx}&". ","")&'
            f'IF(C{row_idx}<>"","Issue date: "&TEXT(C{row_idx},"dd/mm/yyyy")&". ","")&'
            f'IF(D{row_idx}<>"",IF(OR(D{row_idx}="无固定期限",D{row_idx}="长期"),"Valid indefinitely.","Valid until "&D{row_idx}&"."),""))'
        )
    merge_write(ws, "A37:D40", "=TEXTJOIN(CHAR(10),TRUE,J30:J35)", fill=FILL_LABEL, align=Alignment(horizontal="left", vertical="top", wrap_text=True))
    ws.row_dimensions[37].height = 72
    ws.column_dimensions["J"].hidden = True

    merge_write(ws, "A42:E42", "四、年度员工流失率 / Annual Employee Turnover Rate (%)", font=FONT_SECTION, fill=FILL_SECTION, align=Alignment(horizontal="left", vertical="center"))
    merge_write(ws, "A43:E43", "统计口径：过去12个月内离职员工数占全年现场平均员工总数的百分比。输入格式为百分比数值，例如 6.7 表示 6.7%。", fill=FILL_NOTE, align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    write_row(ws, 44, ["时间段", "男（%）", "女（%）", "合计（%）", "备注"], [{"font": FONT_BOLD, "fill": FILL_TABLE_HEAD}] * 5)
    turnover = {
        45: ["上季度（近90天）", 0, 6.7, "=B45+C45"],
        46: ["2025年（全年）", 6.7, 0, "=B46+C46"],
        47: ["2024年（全年）", 0, 6.7, "=B47+C47"],
    }
    for row_idx, row_data in turnover.items():
        style_cell(ws.cell(row=row_idx, column=1, value=row_data[0]), fill=FILL_LABEL, align=Alignment(horizontal="left", vertical="center"))
        style_cell(ws.cell(row=row_idx, column=2, value=row_data[1]), fill=FILL_WHITE, number_format='0.0"%"')
        style_cell(ws.cell(row=row_idx, column=3, value=row_data[2]), fill=FILL_WHITE, number_format='0.0"%"')
        style_cell(ws.cell(row=row_idx, column=4, value=row_data[3]), fill=FILL_CALC, number_format='0.0"%"')
        style_cell(ws.cell(row=row_idx, column=5, value=""), fill=FILL_WHITE)

    merge_write(ws, "A49:C49", "五、环境数据 / Environmental Data", font=FONT_SECTION, fill=FILL_SECTION, align=Alignment(horizontal="left", vertical="center"))
    merge_write(ws, "A50:C50", "注：4P 成员需填写；2P 成员不需要填写本节。", fill=FILL_NOTE, align=Alignment(horizontal="left", vertical="center"))
    write_row(ws, 51, ["指标", "2025年", "2024年"], [{"font": FONT_BOLD, "fill": FILL_TABLE_HEAD}] * 3)
    env_rows = {
        52: ["非可再生能源用电量（kWh）", 44417, 44026],
        53: ["可再生能源用电量（kWh）", 0, 0],
        54: ["总用电量（自动合计）", "=B52+B53", "=C52+C53"],
        55: ["天然气总消耗量（kWh）", "", ""],
        56: ["用水量（m³）", 61, 60],
        57: ["产生的废物总量（mt）", "", ""],
        58: ["产生的危险废物总量（mt）", 0, 0],
    }
    for row_idx, row_data in env_rows.items():
        style_cell(ws.cell(row=row_idx, column=1, value=row_data[0]), fill=FILL_LABEL, align=Alignment(horizontal="left", vertical="center"))
        fill_b = FILL_CALC if row_idx == 54 else FILL_WHITE
        fill_c = FILL_CALC if row_idx == 54 else FILL_WHITE
        style_cell(ws.cell(row=row_idx, column=2, value=row_data[1]), fill=fill_b)
        style_cell(ws.cell(row=row_idx, column=3, value=row_data[2]), fill=fill_c)

    merge_write(ws, "A60:E60", "六、工资信息 / Wage Information", font=FONT_SECTION, fill=FILL_SECTION, align=Alignment(horizontal="left", vertical="center"))
    merge_write(ws, "A61:E61", "计算依据：月薪 ÷ 21.75工作日 ÷ 8小时 = 时薪；加班费时薪 = 时薪 × 1.5。仅需填入每月工资，其余自动计算。", fill=FILL_NOTE, align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    write_row(ws, 62, ["类别", "每小时（元）", "每天（元）", "每周（元）", "每月（元）"], [{"font": FONT_BOLD, "fill": FILL_TABLE_HEAD}] * 5)

    wage_rows = {
        63: ["法定最低工资", "=E63/21.75/8", "=B63*8", "=C63*5", 2500],
        64: ["法定最低加班费（时薪 × 1.5）", "=B63*1.5", "", "", ""],
        66: ["工厂实际给的最低工资", "=E66/21.75/8", "=B66*8", "=C66*5", 2500],
        67: ["工厂实际最低加班费（时薪 × 1.5）", "=B66*1.5", "", "", ""],
        69: ["法定 vs 实际对比", diff_formula("B66", "B63", 2), diff_formula("C66", "C63", 2), diff_formula("D66", "D63", 2), diff_formula("E66", "E63", 0)],
    }

    for row_idx, row_data in wage_rows.items():
        fill = FILL_LABEL if row_idx not in (69,) else FILL_LABEL
        font = FONT_BOLD if row_idx == 69 else FONT_NORMAL
        style_cell(ws.cell(row=row_idx, column=1, value=row_data[0]), fill=fill, font=font, align=Alignment(horizontal="left", vertical="center"))
        for col_idx, value in enumerate(row_data[1:], start=2):
            cell_fill = FILL_GREEN if row_idx == 69 else (FILL_CALC if col_idx < 5 or row_idx in (64, 67) else FILL_WHITE)
            cell_font = FONT_GREEN if row_idx == 69 else FONT_NORMAL
            style_cell(ws.cell(row=row_idx, column=col_idx, value=value), fill=cell_fill, font=cell_font)
        if row_idx in (63, 66):
            ws.cell(row=row_idx, column=2).number_format = "0.00"
            ws.cell(row=row_idx, column=3).number_format = "0.00"
            ws.cell(row=row_idx, column=4).number_format = "0.00"
            ws.cell(row=row_idx, column=5).number_format = "0"
        if row_idx in (64, 67):
            ws.cell(row=row_idx, column=2).number_format = "0.00"
            merge_write(ws, f"C{row_idx}:E{row_idx}", "日/周加班费请按实际加班时数另行计算", fill=FILL_LABEL, font=FONT_SUBTITLE)

    merge_write(ws, "A71:D71", "七、厂房信息 / Factory Building Information", font=FONT_SECTION, fill=FILL_SECTION, align=Alignment(horizontal="left", vertical="center"))
    merge_write(ws, "A72:D72", "厂房 1 / Building 1", font=FONT_BOLD, fill=FILL_TABLE_HEAD, align=Alignment(horizontal="left", vertical="center"))
    building_one = {
        73: ["哪一年竣工？（Year of completion）", 2010],
        74: ["总建筑面积（Total floor area）", "工厂只用400平米"],
        75: ["楼层布局（每层用途 / 生产流程）", "cutting, sewing, ironing, inspection and packing"],
    }
    for row_idx, row_data in building_one.items():
        style_cell(ws.cell(row=row_idx, column=1, value=row_data[0]), fill=FILL_LABEL, align=Alignment(horizontal="left", vertical="center"))
        merge_write(ws, f"B{row_idx}:D{row_idx}", row_data[1], align=Alignment(horizontal="left", vertical="center", wrap_text=True))

    merge_write(ws, "A77:D77", "厂房 2 / Building 2（如有）", font=FONT_BOLD, fill=FILL_TABLE_HEAD, align=Alignment(horizontal="left", vertical="center"))
    building_two = {
        78: ["哪一年竣工？", ""],
        79: ["总建筑面积", ""],
        80: ["楼层布局", ""],
    }
    for row_idx, row_data in building_two.items():
        style_cell(ws.cell(row=row_idx, column=1, value=row_data[0]), fill=FILL_LABEL, align=Alignment(horizontal="left", vertical="center"))
        merge_write(ws, f"B{row_idx}:D{row_idx}", row_data[1], align=Alignment(horizontal="left", vertical="center", wrap_text=True))

    merge_write(ws, "A82:D82", "如有更多楼栋，请按上方格式继续填写。", fill=FILL_NOTICE, align=Alignment(horizontal="left", vertical="center"))
    ws.print_area = "A1:H82"

    for row_idx in range(1, 83):
        if ws.row_dimensions[row_idx].height is None:
            ws.row_dimensions[row_idx].height = 20

    output_path = "sedex审核信息收集表.xlsx"
    wb.save(output_path)
    print(f"created {output_path}")


if __name__ == "__main__":
    main()